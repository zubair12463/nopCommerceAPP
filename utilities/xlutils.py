import openpyxl
from openpyxl.styles import PatternFill

def GetRowCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_row)

def GetCoumnCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_column)

def writedata(file,sheeName,Row,Column,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheeName]
    sheet.cell(Row,Column).value = data
    workbook.save(file)

def readData(file,sheetName,Row,Column):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.cell(Row,Column).value)

def fillGreenColur(file,sheetName,Row,Column):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    greenFill = PatternFill(start_color='60b212',
                            end_color='60b212',
                            fill_type='solid')
    sheet.cell(Row,Column).fill = greenFill
    workbook.save(file)

def fillRedColur(file,sheetName,Row,Column):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]

    redFill = PatternFill(start_color='ff0000',
                          end_color='ff0000',
                          fill_type='solid')
    
    sheet.cell(Row,Column).fill = redFill
    workbook.save(file)